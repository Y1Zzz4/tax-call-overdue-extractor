from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from tax_call_overdue_extractor.config import (
    ExcelSettings,
    LLMReservedSettings,
    LLMSettings,
    LoggingSettings,
    PathSettings,
    ProjectSettings,
    SamplingSettings,
)
from tax_call_overdue_extractor.excel_io import EXPECTED_COLUMNS
from tax_call_overdue_extractor.exceptions import ExtractionError, LLMClientError
from tax_call_overdue_extractor.extraction.batch_service import BatchExtractionService, BatchOptions
from tax_call_overdue_extractor.extraction.normalization import normalize_extraction_result, parse_reference_month
from tax_call_overdue_extractor.extraction.parser import parse_extraction_response
from tax_call_overdue_extractor.llm.client import LLMResponse


def make_settings(tmp_path: Path, *, max_input_chars: int = 12000) -> ProjectSettings:
    return ProjectSettings(
        paths=PathSettings(
            input_dir=tmp_path / "data" / "input",
            samples_dir=tmp_path / "data" / "samples",
            output_dir=tmp_path / "data" / "output",
            conflicts_dir=tmp_path / "data" / "conflicts",
            state_dir=tmp_path / "data" / "state",
            logs_dir=tmp_path / "data" / "logs",
        ),
        excel=ExcelSettings(header_row=1, sheet_name=None, use_active_sheet_when_sheet_not_set=True),
        sampling=SamplingSettings(default_sample_size=50, default_seed=None),
        logging=LoggingSettings(level="INFO"),
        llm_reserved=LLMReservedSettings(
            interface="openai_compatible",
            max_concurrency=2,
            max_retries=3,
            timeout_seconds=60,
            max_input_chars=max_input_chars,
        ),
        llm=LLMSettings(
            base_url="https://example.test/v1",
            api_key="test-key",
            model="mock-model",
            max_retries=3,
            max_input_chars=max_input_chars,
        ),
    )


def create_batch_workbook(path: Path, rows: int, *, blank_rows: set[int] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    blank_rows = blank_rows or set()
    wb = Workbook()
    ws = wb.active
    ws.title = "样本"
    ws.append(list(EXPECTED_COLUMNS))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
    for i in range(1, rows + 1):
        if i in blank_rows:
            voice, business, reply = None, "   ", "#N/A"
        else:
            voice, business, reply = None, f"企业{i} 2025年5月 增值税", None
        ws.append(
            [
                i,
                f"BIZ-{i:04d}",
                f"1380000{i:04d}",
                f"登记人{i}",
                voice,
                "咨询",
                "部门",
                "2026-06-15",
                "2026-06",
                business,
                reply,
                None,
                None,
                None,
                None,
                None,
            ]
        )
        for cell in ws[i + 1]:
            cell.font = Font(name="Arial", size=10)
        ws.row_dimensions[i + 1].height = 23
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:P{ws.max_row}"
    wb.save(path)
    wb.close()


def extraction_json(
    *,
    enterprise: str | None = "企业1",
    tax_types: list[str] | None = None,
    periods: list[dict] | None = None,
    explicitly_overdue: bool | None = None,
    needs_review: bool = False,
    conflicts: list[dict] | None = None,
    amounts: list[dict] | None = None,
) -> str:
    item = {
        "enterprise_name": enterprise,
        "enterprise_evidence": ([{"source": "业务内容", "quote": enterprise}] if enterprise else []),
        "tax_types": tax_types or [],
        "tax_type_raw": tax_types or [],
        "tax_evidence": ([{"source": "业务内容", "quote": tax_types[0]}] if tax_types else []),
        "periods": periods or [],
        "amounts": amounts or [],
        "explicitly_overdue": explicitly_overdue,
        "overdue_evidence": ([{"source": "业务内容", "quote": "已逾期"}] if explicitly_overdue else []),
        "relationship_note": None,
        "needs_review": needs_review,
        "review_reasons": [],
    }
    return json.dumps(
        {
            "schema_version": "1.0",
            "has_relevant_information": True,
            "items": [item],
            "conflicts": conflicts or [],
            "needs_review": needs_review or bool(conflicts),
            "review_reasons": [],
        },
        ensure_ascii=False,
    )


def period(year: int, month: int) -> dict:
    return {
        "raw_text": f"{year}年{month}月",
        "period_type": "single_month",
        "start_year": year,
        "start_month": month,
        "end_year": year,
        "end_month": month,
        "relative_expression": None,
        "evidence": [{"source": "业务内容", "quote": f"{year}年{month}月"}],
    }


class RoutingClient:
    def __init__(self, *, mode: str = "normal") -> None:
        self.mode = mode
        self.calls: list[dict] = []

    def complete(self, request):
        data = json.loads(request.messages[1]["content"])
        assert set(data.keys()) == {"电话录音转文本内容", "业务内容", "答复内容"}
        self.calls.append(data)
        business = data["业务内容"] or ""
        if self.mode == "fail_one" and "企业2" in business:
            raise FakeLLMError("sanitized api failure")
        if self.mode == "multi" and "企业1" in business:
            return LLMResponse(content=json.dumps({
                "schema_version": "1.0",
                "has_relevant_information": True,
                "items": [
                    json.loads(extraction_json(enterprise="甲企业", tax_types=["增值税"], periods=[period(2025, 5)]) )["items"][0],
                    json.loads(extraction_json(enterprise="乙企业", tax_types=["企业所得税"], periods=[period(2025, 6)]) )["items"][0],
                ],
                "conflicts": [],
                "needs_review": False,
                "review_reasons": [],
            }, ensure_ascii=False))
        return LLMResponse(
            content=extraction_json(
                enterprise=business.split()[0] if business else None,
                tax_types=["增值税", "增值税", "城市建设维护税"],
                periods=[period(2025, 5)],
            )
        )


class FakeLLMError(LLMClientError):
    pass


def test_50_rows_batch_processing_and_outputs(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.samples_dir / "sample.xlsx"
    create_batch_workbook(source, 50)
    client = RoutingClient()

    summary = BatchExtractionService(settings, llm_client=client).run(
        BatchOptions(input_path=source, execute=True, overwrite=True, resume=True)
    )

    assert len(client.calls) == 50
    assert summary.success_count == 50
    assert summary.output_path.exists()
    assert summary.conflicts_output_path.exists()
    assert summary.review_output_path.exists()
    wb = load_workbook(summary.output_path)
    try:
        ws = wb.active
        assert ws.max_row == 51
        assert ws.cell(row=2, column=12).value == "企业1"
        assert ws.cell(row=2, column=13).value == "增值税；城市建设维护税"
        assert ws.cell(row=2, column=14).value == "2025年5月到2025年5月"
        assert ws.cell(row=2, column=16).value == "已逾期"
        assert ws.cell(row=2, column=17).value
        assert ws.cell(row=1, column=17).value == "说明"
        assert ws.cell(row=2, column=1).font.name == "Arial"
        assert ws.freeze_panes == "A2"
    finally:
        wb.close()


def test_single_allowed_field_calls_model_and_blank_row_skips(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.samples_dir / "sample.xlsx"
    create_batch_workbook(source, 3, blank_rows={2})
    client = RoutingClient()

    summary = BatchExtractionService(settings, llm_client=client).run(
        BatchOptions(input_path=source, execute=True, overwrite=True)
    )

    assert len(client.calls) == 2
    assert summary.skipped_count == 1
    wb = load_workbook(summary.output_path)
    try:
        ws = wb.active
        assert ws.cell(row=3, column=12).value == "未识别"
        assert ws.cell(row=3, column=13).value == "未识别"
        assert ws.cell(row=3, column=14).value == "未识别"
        assert ws.cell(row=3, column=15).value == "未提及"
        assert ws.cell(row=3, column=16).value == "未明确"
        assert ws.cell(row=3, column=17).value == "三列均无可分析文本"
    finally:
        wb.close()


def test_single_failure_does_not_stop_batch(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.samples_dir / "sample.xlsx"
    create_batch_workbook(source, 3)
    client = RoutingClient(mode="fail_one")

    summary = BatchExtractionService(settings, llm_client=client).run(
        BatchOptions(input_path=source, execute=True, overwrite=True)
    )

    assert summary.api_error_count == 1
    assert summary.success_count == 2
    assert summary.review_output_path.exists()


def test_resume_reuses_success_and_reprocesses_changed_input_or_prompt(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.samples_dir / "sample.xlsx"
    create_batch_workbook(source, 2)
    first_client = RoutingClient()
    service = BatchExtractionService(settings, llm_client=first_client, system_prompt="prompt-v1")
    service.run(BatchOptions(input_path=source, execute=True, overwrite=True, resume=True))
    assert len(first_client.calls) == 2

    second_client = RoutingClient()
    BatchExtractionService(settings, llm_client=second_client, system_prompt="prompt-v1").run(
        BatchOptions(input_path=source, execute=True, overwrite=True, resume=True)
    )
    assert len(second_client.calls) == 0

    third_client = RoutingClient()
    BatchExtractionService(settings, llm_client=third_client, system_prompt="prompt-v2").run(
        BatchOptions(input_path=source, execute=True, overwrite=True, resume=True)
    )
    assert len(third_client.calls) == 2

    wb = load_workbook(source)
    try:
        wb.active.cell(row=2, column=10).value = "变更企业 2025年5月 增值税"
        wb.save(source)
    finally:
        wb.close()
    fourth_client = RoutingClient()
    BatchExtractionService(settings, llm_client=fourth_client, system_prompt="prompt-v2").run(
        BatchOptions(input_path=source, execute=True, overwrite=True, resume=True)
    )
    assert len(fourth_client.calls) == 1


def test_multi_item_insert_rows_and_first_11_columns_blank(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.samples_dir / "sample.xlsx"
    create_batch_workbook(source, 2)
    client = RoutingClient(mode="multi")

    summary = BatchExtractionService(settings, llm_client=client).run(
        BatchOptions(input_path=source, rows=(2,), execute=True, overwrite=True)
    )

    wb = load_workbook(summary.output_path)
    try:
        ws = wb.active
        assert ws.max_row == 4
        assert ws.cell(row=2, column=12).value == "甲企业"
        assert ws.cell(row=3, column=12).value == "乙企业"
        assert all(ws.cell(row=3, column=col).value is None for col in range(1, 12))
        assert ws.cell(row=3, column=1).font.name == ws.cell(row=2, column=1).font.name
    finally:
        wb.close()


def test_period_normalization_and_overdue_rules() -> None:
    reference = parse_reference_month("2026-01-12", "2026-01")
    result = parse_extraction_response(extraction_json(periods=[
        {
            "raw_text": "上个月",
            "period_type": "relative",
            "start_year": None,
            "start_month": None,
            "end_year": None,
            "end_month": None,
            "relative_expression": "上个月",
            "evidence": [{"source": "业务内容", "quote": "上个月"}],
        }
    ]))
    item = normalize_extraction_result(result, reference_month=reference)[0]
    assert item.periods_text == "2025年12月到2025年12月"
    assert item.overdue_text == "已逾期"

    equal = normalize_extraction_result(parse_extraction_response(extraction_json(periods=[period(2026, 1)])), reference_month=reference)[0]
    future = normalize_extraction_result(parse_extraction_response(extraction_json(periods=[period(2026, 2)])), reference_month=reference)[0]
    year = normalize_extraction_result(parse_extraction_response(extraction_json(periods=[{
        "raw_text": "2025年",
        "period_type": "year",
        "start_year": 2025,
        "start_month": None,
        "end_year": None,
        "end_month": None,
        "relative_expression": None,
        "evidence": [{"source": "业务内容", "quote": "2025年"}],
    }])), reference_month=reference)[0]
    assert equal.overdue_text is None
    assert future.overdue_text is None
    assert year.periods_text == "2025年1月到2025年12月"
    assert year.overdue_text is None


def test_explicit_overdue_and_conflicting_status_outputs(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.samples_dir / "sample.xlsx"
    create_batch_workbook(source, 1)

    class ConflictClient(RoutingClient):
        def complete(self, request):
            self.calls.append(json.loads(request.messages[1]["content"]))
            return LLMResponse(content=extraction_json(explicitly_overdue=False, periods=[period(2025, 5)]))

    summary = BatchExtractionService(settings, llm_client=ConflictClient()).run(
        BatchOptions(input_path=source, execute=True, overwrite=True)
    )

    assert summary.conflict_count == 0
    assert summary.success_count == 1
    wb = load_workbook(summary.output_path)
    try:
        assert wb.active.cell(row=2, column=16).value == "未逾期"
    finally:
        wb.close()
    conflict_wb = load_workbook(summary.conflicts_output_path)
    try:
        assert conflict_wb.active.max_row == 1
    finally:
        conflict_wb.close()


def test_preflight_does_not_call_api_or_expose_text(tmp_path: Path, capsys) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.samples_dir / "sample.xlsx"
    create_batch_workbook(source, 3)
    plan = BatchExtractionService(settings, llm_client=RoutingClient()).preflight(BatchOptions(input_path=source))

    assert plan.estimated_api_calls == 3
    assert plan.eligible_rows == 3
    assert "企业1" not in capsys.readouterr().out


def test_safety_limit_blocks_more_than_100_records(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.samples_dir / "sample.xlsx"
    create_batch_workbook(source, 101)

    with pytest.raises(ExtractionError, match="安全上限"):
        BatchExtractionService(settings, llm_client=RoutingClient()).run(
            BatchOptions(input_path=source, execute=True, overwrite=True)
        )
