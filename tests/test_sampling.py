from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from tax_call_overdue_extractor.config import (
    ExcelSettings,
    LLMReservedSettings,
    LoggingSettings,
    PathSettings,
    ProjectSettings,
    SamplingSettings,
)
from tax_call_overdue_extractor.excel_io import EXPECTED_COLUMNS, hash_file
from tax_call_overdue_extractor.exceptions import ExcelIOError, SamplingError
from tax_call_overdue_extractor.sampling import is_valid_call_text, sample_excel_file


def make_settings(tmp_path: Path) -> ProjectSettings:
    return ProjectSettings(
        paths=PathSettings(
            input_dir=tmp_path / "data" / "input",
            samples_dir=tmp_path / "data" / "samples",
            output_dir=tmp_path / "data" / "output",
            conflicts_dir=tmp_path / "data" / "conflicts",
            state_dir=tmp_path / "data" / "state",
            logs_dir=tmp_path / "data" / "logs",
        ),
        excel=ExcelSettings(
            header_row=1,
            sheet_name=None,
            use_active_sheet_when_sheet_not_set=True,
        ),
        sampling=SamplingSettings(default_sample_size=50, default_seed=None),
        logging=LoggingSettings(level="INFO"),
        llm_reserved=LLMReservedSettings(
            interface="openai_compatible",
            max_concurrency=4,
            max_retries=3,
            timeout_seconds=60,
        ),
    )


def create_source_workbook(
    path: Path,
    *,
    valid_rows: int,
    invalid_values: list[object] | None = None,
    blank_business_rows: set[int] | None = None,
    blank_answer_rows: set[int] | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    invalid_values = invalid_values or []
    blank_business_rows = blank_business_rows or set()
    blank_answer_rows = blank_answer_rows or set()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "税务电话"
    worksheet.append(list(EXPECTED_COLUMNS))

    header_fill = PatternFill("solid", fgColor="FFF2CC")
    text_fill = PatternFill("solid", fgColor="D9EAD3")
    thin_side = Side(style="thin", color="666666")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for cell in worksheet[1]:
        cell.font = Font(name="Microsoft YaHei", bold=True, color="1F4E78", size=11)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 26

    all_call_values = [f"有效电话文本 {index}" for index in range(1, valid_rows + 1)]
    all_call_values.extend(invalid_values)
    for index, call_text in enumerate(all_call_values, start=1):
        business_value = None if index in blank_business_rows else f"业务内容 {index}"
        answer_value = "#N/A" if index in blank_answer_rows else f"答复内容 {index}"
        worksheet.append(
            [
                index,
                f"BIZ-{index:04d}",
                f"1380000{index:04d}",
                f"登记人{index}",
                call_text,
                "咨询",
                "第一税务所",
                datetime(2026, 1, (index % 28) + 1),
                "2026-01",
                business_value,
                answer_value,
                f"测试企业{index}",
                None,
                None,
                None,
                None,
            ]
        )
        row = index + 1
        worksheet.row_dimensions[row].height = 24
        for column in range(1, len(EXPECTED_COLUMNS) + 1):
            cell = worksheet.cell(row=row, column=column)
            cell.font = Font(name="Arial", size=10, color="000000")
            cell.fill = text_fill if column == 5 else PatternFill(fill_type=None)
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        worksheet.cell(row=row, column=8).number_format = "yyyy-mm-dd"

    for column in range(1, len(EXPECTED_COLUMNS) + 1):
        letter = worksheet.cell(row=1, column=column).column_letter
        worksheet.column_dimensions[letter].width = 18 + column
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:P{worksheet.max_row}"

    table = Table(displayName="SourceTable", ref=f"A1:P{worksheet.max_row}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    worksheet.add_table(table)
    workbook.save(path)
    workbook.close()


def read_column_values(path: Path, column: int) -> list[object]:
    workbook = load_workbook(path)
    try:
        worksheet = workbook.active
        return [worksheet.cell(row=row, column=column).value for row in range(2, worksheet.max_row + 1)]
    finally:
        workbook.close()


def test_extracts_50_rows_normally(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.input_dir / "source.xlsx"
    output = settings.paths.samples_dir / "source_sample_50.xlsx"
    create_source_workbook(source, valid_rows=80)

    result = sample_excel_file(
        settings=settings,
        input_path=source,
        output_path=output,
        sample_size=50,
        seed=2026,
    )

    workbook = load_workbook(output)
    try:
        worksheet = workbook.active
        assert result.candidate_count == 80
        assert result.sample_size == 50
        assert worksheet.max_row == 51
        assert worksheet.max_column == 16
        assert [worksheet.cell(row=1, column=i).value for i in range(1, 17)] == list(EXPECTED_COLUMNS)
    finally:
        workbook.close()


def test_invalid_call_text_values_are_excluded(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.input_dir / "source.xlsx"
    output = settings.paths.samples_dir / "sample.xlsx"
    create_source_workbook(source, valid_rows=10, invalid_values=["#N/A", "  #n/a  ", None, "   "])

    result = sample_excel_file(
        settings=settings,
        input_path=source,
        output_path=output,
        sample_size=10,
        seed=7,
    )

    assert result.candidate_count == 10
    assert all(is_valid_call_text(value) for value in read_column_values(output, 5))


def test_validity_helper_handles_na_case_space_and_blank_values() -> None:
    assert not is_valid_call_text("#N/A")
    assert not is_valid_call_text("  #n/a  ")
    assert not is_valid_call_text(None)
    assert not is_valid_call_text("   ")
    assert is_valid_call_text("纳税人咨询逾期事项")


def test_blank_business_or_answer_does_not_affect_sampling(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.input_dir / "source.xlsx"
    output = settings.paths.samples_dir / "sample.xlsx"
    create_source_workbook(
        source,
        valid_rows=20,
        blank_business_rows={1, 2, 3},
        blank_answer_rows={4, 5, 6},
    )

    result = sample_excel_file(
        settings=settings,
        input_path=source,
        output_path=output,
        sample_size=20,
        seed=3,
    )

    assert result.candidate_count == 20
    workbook = load_workbook(output)
    try:
        worksheet = workbook.active
        assert worksheet.cell(row=2, column=10).value is None
        assert "#N/A" in [worksheet.cell(row=row, column=11).value for row in range(2, 22)]
    finally:
        workbook.close()


def test_same_seed_produces_same_rows(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.input_dir / "source.xlsx"
    create_source_workbook(source, valid_rows=100)

    first = sample_excel_file(
        settings=settings,
        input_path=source,
        output_path=settings.paths.samples_dir / "first.xlsx",
        sample_size=20,
        seed=2026,
    )
    second = sample_excel_file(
        settings=settings,
        input_path=source,
        output_path=settings.paths.samples_dir / "second.xlsx",
        sample_size=20,
        seed=2026,
    )

    assert first.selected_rows == second.selected_rows


def test_different_seed_can_produce_different_rows(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.input_dir / "source.xlsx"
    create_source_workbook(source, valid_rows=100)

    first = sample_excel_file(
        settings=settings,
        input_path=source,
        output_path=settings.paths.samples_dir / "first.xlsx",
        sample_size=20,
        seed=1,
    )
    second = sample_excel_file(
        settings=settings,
        input_path=source,
        output_path=settings.paths.samples_dir / "second.xlsx",
        sample_size=20,
        seed=2,
    )

    assert first.selected_rows != second.selected_rows


def test_output_rows_keep_source_order(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.input_dir / "source.xlsx"
    output = settings.paths.samples_dir / "sample.xlsx"
    create_source_workbook(source, valid_rows=90)

    result = sample_excel_file(
        settings=settings,
        input_path=source,
        output_path=output,
        sample_size=30,
        seed=99,
    )

    output_sequence_numbers = read_column_values(output, 1)
    expected_sequence_numbers = [row - 1 for row in result.selected_rows]
    assert output_sequence_numbers == expected_sequence_numbers
    assert output_sequence_numbers == sorted(output_sequence_numbers)


def test_header_columns_values_and_styles_are_preserved(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.input_dir / "source.xlsx"
    output = settings.paths.samples_dir / "sample.xlsx"
    create_source_workbook(source, valid_rows=75)

    result = sample_excel_file(
        settings=settings,
        input_path=source,
        output_path=output,
        sample_size=25,
        seed=11,
    )

    source_workbook = load_workbook(source)
    output_workbook = load_workbook(output)
    try:
        source_sheet = source_workbook.active
        output_sheet = output_workbook.active
        selected_source_row = result.selected_rows[0]

        assert output_sheet.max_column == 16
        assert [output_sheet.cell(row=1, column=i).value for i in range(1, 17)] == list(EXPECTED_COLUMNS)
        assert output_sheet.cell(row=2, column=1).value == source_sheet.cell(row=selected_source_row, column=1).value
        assert output_sheet.cell(row=2, column=5).fill.fgColor.rgb == source_sheet.cell(
            row=selected_source_row, column=5
        ).fill.fgColor.rgb
        assert output_sheet.cell(row=2, column=5).font.name == "Arial"
        assert output_sheet.cell(row=2, column=5).border.left.style == "thin"
        assert output_sheet.cell(row=2, column=5).alignment.wrap_text is True
        assert output_sheet.cell(row=2, column=8).number_format == "yyyy-mm-dd"
        assert output_sheet.row_dimensions[2].height == source_sheet.row_dimensions[selected_source_row].height
        assert output_sheet.column_dimensions["E"].width == source_sheet.column_dimensions["E"].width
        assert output_sheet.freeze_panes == source_sheet.freeze_panes
        assert output_sheet.auto_filter.ref == "A1:P26"
        assert output_sheet.tables["SourceTable"].ref == "A1:P26"
    finally:
        source_workbook.close()
        output_workbook.close()


def test_insufficient_candidates_raise_error_and_no_output(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.input_dir / "source.xlsx"
    output = settings.paths.samples_dir / "sample.xlsx"
    create_source_workbook(source, valid_rows=5)

    with pytest.raises(SamplingError, match="有效数据不足"):
        sample_excel_file(
            settings=settings,
            input_path=source,
            output_path=output,
            sample_size=50,
            seed=1,
        )

    assert not output.exists()


def test_multiple_input_files_require_explicit_input(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    create_source_workbook(settings.paths.input_dir / "a.xlsx", valid_rows=60)
    create_source_workbook(settings.paths.input_dir / "b.xlsx", valid_rows=60)

    with pytest.raises(ExcelIOError, match="多个 .xlsx 文件"):
        sample_excel_file(settings=settings, sample_size=50, seed=1)


def test_existing_output_is_not_overwritten_by_default(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.input_dir / "source.xlsx"
    output = settings.paths.samples_dir / "sample.xlsx"
    create_source_workbook(source, valid_rows=60)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text("existing", encoding="utf-8")

    with pytest.raises(ExcelIOError, match="默认拒绝覆盖"):
        sample_excel_file(
            settings=settings,
            input_path=source,
            output_path=output,
            sample_size=50,
            seed=1,
        )

    assert output.read_text(encoding="utf-8") == "existing"


def test_input_file_is_not_modified(tmp_path: Path) -> None:
    settings = make_settings(tmp_path)
    source = settings.paths.input_dir / "source.xlsx"
    output = settings.paths.samples_dir / "sample.xlsx"
    create_source_workbook(source, valid_rows=70)
    before = hash_file(source)

    sample_excel_file(
        settings=settings,
        input_path=source,
        output_path=output,
        sample_size=50,
        seed=123,
    )

    assert hash_file(source) == before
