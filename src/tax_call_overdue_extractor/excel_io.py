"""Excel 文件检查、复制、删行和输出验证。"""

from __future__ import annotations

import hashlib
import os
import shutil
import uuid
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from .exceptions import (
    ExcelIOError,
    HeaderValidationError,
    OutputValidationError,
    SheetNotFoundError,
)
from .sampling import is_valid_call_text


EXPECTED_COLUMNS: tuple[str, ...] = (
    "序号",
    "业务编号",
    "来电号码",
    "登记人姓名",
    "语音转文本",
    "小结类型",
    "登记人部门名称",
    "登记日期",
    "月份",
    "业务内容",
    "答复内容",
    "企业名称",
    "逾期税种",
    "所属期",
    "涉及金额",
    "是否确定已逾期",
)
CALL_TEXT_COLUMN = "语音转文本"


@dataclass(frozen=True)
class SourceWorkbookInfo:
    """源工作簿的非敏感结构信息。"""

    path: Path
    sheet_name: str
    header: tuple[str, ...]
    header_row: int
    text_column_index: int
    candidate_rows: tuple[int, ...]


@dataclass(frozen=True)
class RowDimensionSnapshot:
    height: float | None
    hidden: bool
    outline_level: int
    collapsed: bool
    thick_top: bool
    thick_bottom: bool
    style: object | None


def resolve_input_file(input_path: str | Path | None, input_dir: Path) -> Path:
    """解析输入文件；未指定时要求 input_dir 下恰好有一个 .xlsx 文件。"""

    if input_path is not None:
        path = Path(input_path)
        _validate_input_xlsx(path)
        return path

    if not input_dir.exists() or not input_dir.is_dir():
        raise ExcelIOError(f"输入目录不存在或不可访问: {input_dir}")

    files = sorted(
        path for path in input_dir.glob("*.xlsx") if path.is_file() and not path.name.startswith("~$")
    )
    if not files:
        raise ExcelIOError(f"输入目录下没有 .xlsx 文件: {input_dir}")
    if len(files) > 1:
        raise ExcelIOError(f"输入目录下存在多个 .xlsx 文件，请使用 --input 指定输入文件: {input_dir}")
    _validate_input_xlsx(files[0])
    return files[0]


def build_default_sample_output_path(input_path: Path, samples_dir: Path, sample_size: int) -> Path:
    """根据输入文件名生成默认抽样输出路径。"""

    return samples_dir / f"{input_path.stem}_sample_{sample_size}.xlsx"


def inspect_source_workbook(
    input_path: Path,
    *,
    sheet_name: str | None,
    use_active_sheet: bool,
    header_row: int,
) -> SourceWorkbookInfo:
    """读取源工作簿结构，返回候选数据行号，不返回任何敏感单元格文本。"""

    _validate_input_xlsx(input_path)
    workbook = _load_workbook(input_path, "读取源工作簿")
    try:
        worksheet = _select_worksheet(workbook, sheet_name, use_active_sheet)
        header = validate_header(worksheet, header_row)
        text_column_index = header.index(CALL_TEXT_COLUMN) + 1
        candidate_rows = tuple(
            row_index
            for row_index in range(header_row + 1, worksheet.max_row + 1)
            if is_valid_call_text(worksheet.cell(row=row_index, column=text_column_index).value)
        )
        return SourceWorkbookInfo(
            path=input_path,
            sheet_name=worksheet.title,
            header=header,
            header_row=header_row,
            text_column_index=text_column_index,
            candidate_rows=candidate_rows,
        )
    finally:
        workbook.close()


def create_sample_workbook(
    *,
    input_path: Path,
    output_path: Path,
    selected_rows: tuple[int, ...],
    sheet_name: str,
    header_row: int,
    use_active_sheet: bool,
    overwrite: bool,
) -> None:
    """复制源 Excel 到临时文件，删除未抽中行，验证成功后移动到输出路径。"""

    _validate_input_xlsx(input_path)
    _validate_output_path(input_path, output_path, overwrite)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    source_hash = hash_file(input_path)
    temp_path = output_path.parent / f".{output_path.stem}.{uuid.uuid4().hex}.tmp.xlsx"

    try:
        shutil.copy2(input_path, temp_path)
        workbook = _load_workbook(temp_path, "读取临时工作簿")
        try:
            worksheet = _select_worksheet(workbook, sheet_name, use_active_sheet)
            header = validate_header(worksheet, header_row)

            selected_set = set(selected_rows)
            header_dimension = _capture_row_dimension(worksheet, header_row)
            selected_dimensions = {
                row_index: _capture_row_dimension(worksheet, row_index) for row_index in selected_rows
            }
            month_column = header.index("月份") + 1
            selected_month_formulas = {
                row_index: worksheet.cell(row=row_index, column=month_column).value
                for row_index in selected_rows
                if worksheet.cell(row=row_index, column=month_column).data_type == "f"
            }

            for row_index in range(worksheet.max_row, header_row, -1):
                if row_index not in selected_set:
                    worksheet.delete_rows(row_index)

            _restore_row_dimensions(worksheet, header_row, header_dimension, selected_rows, selected_dimensions)
            _translate_sample_month_formulas(
                worksheet,
                header_row=header_row,
                month_column=month_column,
                selected_rows=selected_rows,
                source_formulas=selected_month_formulas,
            )
            _update_filter_and_tables(worksheet, header_row, len(selected_rows))
            worksheet.parent.calculation.calcMode = "auto"
            worksheet.parent.calculation.fullCalcOnLoad = True
            worksheet.parent.calculation.forceFullCalc = True
            workbook.save(temp_path)
        finally:
            workbook.close()

        validate_sample_output(
            input_path=input_path,
            output_path=temp_path,
            selected_rows=selected_rows,
            sheet_name=sheet_name,
            header_row=header_row,
            expected_source_hash=source_hash,
        )
        if hash_file(input_path) != source_hash:
            raise OutputValidationError("原始输入文件在处理过程中发生变化，已放弃输出")
        os.replace(temp_path, output_path)
    except (ExcelIOError, OSError, PermissionError, InvalidFileException, BadZipFile) as exc:
        _remove_temp_file(temp_path)
        if isinstance(exc, ExcelIOError):
            raise
        raise ExcelIOError(f"工作簿复制、保存或验证失败: {exc}") from exc
    except Exception as exc:
        _remove_temp_file(temp_path)
        raise ExcelIOError(f"工作簿复制、保存或验证失败: {exc}") from exc


def validate_sample_output(
    *,
    input_path: Path,
    output_path: Path,
    selected_rows: tuple[int, ...],
    sheet_name: str,
    header_row: int,
    expected_source_hash: str | None = None,
) -> None:
    """重新打开输出文件，验证内容、顺序和主要格式均符合抽样结果。"""

    if tuple(selected_rows) != tuple(sorted(selected_rows)):
        raise OutputValidationError("抽样行号未按原文件顺序排序")

    source_workbook = _load_workbook(input_path, "验证源工作簿")
    output_workbook = _load_workbook(output_path, "验证输出工作簿")
    try:
        source_sheet = _select_worksheet(source_workbook, sheet_name, True)
        output_sheet = _select_worksheet(output_workbook, sheet_name, True)
        source_header = validate_header(source_sheet, header_row)
        output_header = validate_header(output_sheet, header_row)
        if output_sheet.title != source_sheet.title:
            raise OutputValidationError("输出工作表名称与源工作表不一致")
        if output_header != source_header:
            raise OutputValidationError("输出表头与源文件不一致")
        if output_sheet.max_column != len(EXPECTED_COLUMNS):
            raise OutputValidationError(f"输出列数不是 16 列: {output_sheet.max_column}")
        expected_max_row = header_row + len(selected_rows)
        if output_sheet.max_row != expected_max_row:
            raise OutputValidationError(
                f"输出数据行数不正确: 实际 {output_sheet.max_row - header_row} 行，要求 {len(selected_rows)} 行"
            )

        text_column = source_header.index(CALL_TEXT_COLUMN) + 1
        for offset, source_row in enumerate(selected_rows, start=1):
            output_row = header_row + offset
            if not is_valid_call_text(output_sheet.cell(row=output_row, column=text_column).value):
                raise OutputValidationError("输出文件包含无效的语音转文本内容")
            for column in range(1, len(EXPECTED_COLUMNS) + 1):
                source_cell = source_sheet.cell(row=source_row, column=column)
                output_cell = output_sheet.cell(row=output_row, column=column)
                if source_header[column - 1] == "月份":
                    _validate_sample_month_cell(source_cell, output_cell)
                    values_equivalent = True
                else:
                    values_equivalent = _cell_values_equivalent(source_cell.value, output_cell.value)
                if not values_equivalent:
                    raise OutputValidationError(
                        _cell_value_mismatch_message(
                            column_name=source_header[column - 1],
                            source_row=source_row,
                            output_row=output_row,
                            source_cell=source_cell,
                            output_cell=output_cell,
                            source_hash_changed=_source_hash_changed(
                                input_path,
                                expected_source_hash,
                            ),
                        )
                    )
                if _cell_style_signature(output_cell) != _cell_style_signature(source_cell):
                    raise OutputValidationError("输出单元格主要样式与源文件抽样行不一致")

        _validate_dimensions_and_sheet_properties(
            source_sheet,
            output_sheet,
            header_row,
            selected_rows,
        )
    finally:
        source_workbook.close()
        output_workbook.close()


def validate_header(worksheet: Worksheet, header_row: int) -> tuple[str, ...]:
    """校验 16 列表头、重复列名和目标文本列。"""

    if worksheet.max_column != len(EXPECTED_COLUMNS):
        raise HeaderValidationError(
            f"表头与预期16列不一致: 实际列数 {worksheet.max_column}，预期 {len(EXPECTED_COLUMNS)}"
        )

    header = tuple(worksheet.cell(row=header_row, column=column).value for column in range(1, 17))
    header_names = tuple("" if value is None else str(value) for value in header)
    duplicates = sorted({name for name in header_names if header_names.count(name) > 1})
    if duplicates:
        raise HeaderValidationError(f"表头存在重复列名: {', '.join(duplicates)}")
    if CALL_TEXT_COLUMN not in header_names:
        raise HeaderValidationError(f"找不到指定列: {CALL_TEXT_COLUMN}")
    if header_names != EXPECTED_COLUMNS:
        raise HeaderValidationError("表头与预期16列不一致")
    return header_names


def hash_file(path: Path) -> str:
    """计算文件哈希，用于确认原始输入未被抽样过程修改。"""

    digest = hashlib.sha256()
    try:
        with path.open("rb") as file:
            for chunk in iter(lambda: file.read(1024 * 1024), b""):
                digest.update(chunk)
    except OSError as exc:
        raise ExcelIOError(f"文件无法访问: {path}") from exc
    return digest.hexdigest()


def _validate_input_xlsx(path: Path) -> None:
    if not path.exists():
        raise ExcelIOError(f"输入文件不存在: {path}")
    if not path.is_file():
        raise ExcelIOError(f"输入路径不是文件: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ExcelIOError(f"输入文件不是 .xlsx 文件: {path}")


def _validate_output_path(input_path: Path, output_path: Path, overwrite: bool) -> None:
    if output_path.suffix.lower() != ".xlsx":
        raise ExcelIOError(f"输出文件必须是 .xlsx 文件: {output_path}")
    if input_path.resolve() == output_path.resolve():
        raise ExcelIOError("输入路径和输出路径不能相同")
    if output_path.exists() and not overwrite:
        raise ExcelIOError(f"输出文件已存在，默认拒绝覆盖: {output_path}")


def _load_workbook(path: Path, action: str):
    try:
        return load_workbook(path)
    except FileNotFoundError as exc:
        raise ExcelIOError(f"{action}失败，文件不存在: {path}") from exc
    except PermissionError as exc:
        raise ExcelIOError(f"{action}失败，文件无权限或被占用: {path}") from exc
    except (InvalidFileException, BadZipFile, OSError) as exc:
        raise ExcelIOError(f"{action}失败，文件可能损坏或不可访问: {path}") from exc


def _select_worksheet(workbook, sheet_name: str | None, use_active_sheet: bool) -> Worksheet:
    if sheet_name is not None:
        if sheet_name not in workbook.sheetnames:
            raise SheetNotFoundError(f"找不到指定工作表: {sheet_name}")
        return workbook[sheet_name]
    if use_active_sheet:
        return workbook.active
    if len(workbook.sheetnames) == 1:
        return workbook[workbook.sheetnames[0]]
    raise SheetNotFoundError("未指定工作表，且工作簿包含多个工作表")


def _capture_row_dimension(worksheet: Worksheet, row_index: int) -> RowDimensionSnapshot:
    dimension = worksheet.row_dimensions[row_index]
    return RowDimensionSnapshot(
        height=dimension.height,
        hidden=dimension.hidden,
        outline_level=dimension.outlineLevel,
        collapsed=dimension.collapsed,
        thick_top=dimension.thickTop,
        thick_bottom=dimension.thickBot,
        style=copy(getattr(dimension, "_style", None)),
    )


def _restore_row_dimensions(
    worksheet: Worksheet,
    header_row: int,
    header_dimension: RowDimensionSnapshot,
    selected_rows: tuple[int, ...],
    selected_dimensions: dict[int, RowDimensionSnapshot],
) -> None:
    _apply_row_dimension(worksheet, header_row, header_dimension)
    for offset, source_row in enumerate(selected_rows, start=1):
        snapshot = selected_dimensions.get(source_row)
        if snapshot is not None:
            _apply_row_dimension(worksheet, header_row + offset, snapshot)


def _translate_sample_month_formulas(
    worksheet: Worksheet,
    *,
    header_row: int,
    month_column: int,
    selected_rows: tuple[int, ...],
    source_formulas: dict[int, object],
) -> None:
    """把抽中行的月份公式从原行坐标平移到样本当前行坐标。"""

    column_letter = get_column_letter(month_column)
    for offset, source_row in enumerate(selected_rows, start=1):
        formula = source_formulas.get(source_row)
        if not isinstance(formula, str):
            continue
        output_row = header_row + offset
        origin = f"{column_letter}{source_row}"
        destination = f"{column_letter}{output_row}"
        worksheet.cell(row=output_row, column=month_column).value = Translator(
            formula,
            origin=origin,
        ).translate_formula(destination)


def _validate_sample_month_cell(source_cell, output_cell) -> None:
    if source_cell.data_type == "f" and isinstance(source_cell.value, str):
        expected_value = Translator(
            source_cell.value,
            origin=source_cell.coordinate,
        ).translate_formula(output_cell.coordinate)
    else:
        expected_value = source_cell.value
    if (
        output_cell.value != expected_value
        or output_cell.data_type != source_cell.data_type
        or output_cell.number_format != source_cell.number_format
    ):
        raise OutputValidationError("抽样月份公式未指向当前行，或月份单元格类型/格式发生变化")


def _apply_row_dimension(
    worksheet: Worksheet,
    row_index: int,
    snapshot: RowDimensionSnapshot,
) -> None:
    dimension = worksheet.row_dimensions[row_index]
    dimension.height = snapshot.height
    dimension.hidden = snapshot.hidden
    dimension.outlineLevel = snapshot.outline_level
    dimension.collapsed = snapshot.collapsed
    dimension.thickTop = snapshot.thick_top
    dimension.thickBot = snapshot.thick_bottom
    if snapshot.style is not None:
        dimension._style = copy(snapshot.style)


def _update_filter_and_tables(worksheet: Worksheet, header_row: int, data_row_count: int) -> None:
    new_ref = _sample_range(header_row, data_row_count)
    if worksheet.auto_filter and worksheet.auto_filter.ref:
        worksheet.auto_filter.ref = new_ref
    for table in worksheet.tables.values():
        table.ref = new_ref
        if table.autoFilter is not None:
            table.autoFilter.ref = new_ref


def _sample_range(header_row: int, data_row_count: int) -> str:
    last_column = get_column_letter(len(EXPECTED_COLUMNS))
    return f"A{header_row}:{last_column}{header_row + data_row_count}"


def _cell_style_signature(cell) -> tuple[object, object, object, object, str]:
    return (
        copy(cell.font),
        copy(cell.fill),
        copy(cell.border),
        copy(cell.alignment),
        cell.number_format,
    )


def _cell_values_equivalent(source_value: object, output_value: object) -> bool:
    """判断保存前后单元格值是否等价；只容忍 Excel 换行符规范化。"""

    if source_value == output_value:
        return True
    if isinstance(source_value, str) and isinstance(output_value, str):
        return _normalize_newlines(source_value) == _normalize_newlines(output_value)
    return False


def _cell_value_mismatch_message(
    *,
    column_name: str,
    source_row: int,
    output_row: int,
    source_cell,
    output_cell,
    source_hash_changed: bool | None,
) -> str:
    """构造不含单元格真实内容的验证诊断信息。"""

    hash_status = "unknown" if source_hash_changed is None else str(source_hash_changed)
    return (
        "输出单元格内容与源文件抽样行不一致 "
        f"column={column_name!r} "
        f"source_row={source_row} "
        f"output_row={output_row} "
        f"source_value_type={type(source_cell.value).__name__} "
        f"output_value_type={type(output_cell.value).__name__} "
        f"source_data_type={source_cell.data_type!r} "
        f"output_data_type={output_cell.data_type!r} "
        f"source_text_length={_text_length(source_cell.value)} "
        f"output_text_length={_text_length(output_cell.value)} "
        f"same_after_strip={_same_after_strip(source_cell.value, output_cell.value)} "
        f"same_after_newline_normalize={_same_after_newline_normalize(source_cell.value, output_cell.value)} "
        f"same_after_whitespace_collapse={_same_after_whitespace_collapse(source_cell.value, output_cell.value)} "
        f"source_hash_changed={hash_status}"
    )


def _source_hash_changed(input_path: Path, expected_source_hash: str | None) -> bool | None:
    if expected_source_hash is None:
        return None
    return hash_file(input_path) != expected_source_hash


def _text_length(value: object) -> int | str:
    return len(value) if isinstance(value, str) else "n/a"


def _same_after_strip(source_value: object, output_value: object) -> bool | str:
    if not isinstance(source_value, str) or not isinstance(output_value, str):
        return "n/a"
    return source_value.strip() == output_value.strip()


def _same_after_newline_normalize(source_value: object, output_value: object) -> bool | str:
    if not isinstance(source_value, str) or not isinstance(output_value, str):
        return "n/a"
    return _normalize_newlines(source_value) == _normalize_newlines(output_value)


def _same_after_whitespace_collapse(source_value: object, output_value: object) -> bool | str:
    if not isinstance(source_value, str) or not isinstance(output_value, str):
        return "n/a"
    return " ".join(source_value.split()) == " ".join(output_value.split())


def _normalize_newlines(value: str) -> str:
    return value.replace("\r\n", "\n").replace("\r", "\n")


def _validate_dimensions_and_sheet_properties(
    source_sheet: Worksheet,
    output_sheet: Worksheet,
    header_row: int,
    selected_rows: tuple[int, ...],
) -> None:
    for column in range(1, len(EXPECTED_COLUMNS) + 1):
        letter = get_column_letter(column)
        if source_sheet.column_dimensions[letter].width != output_sheet.column_dimensions[letter].width:
            raise OutputValidationError("输出列宽与源文件不一致")

    if source_sheet.row_dimensions[header_row].height != output_sheet.row_dimensions[header_row].height:
        raise OutputValidationError("输出表头行高与源文件不一致")
    for offset, source_row in enumerate(selected_rows, start=1):
        output_row = header_row + offset
        if source_sheet.row_dimensions[source_row].height != output_sheet.row_dimensions[output_row].height:
            raise OutputValidationError("输出数据行高与源文件抽样行不一致")

    if source_sheet.freeze_panes != output_sheet.freeze_panes:
        raise OutputValidationError("输出冻结窗格设置与源文件不一致")

    expected_ref = _sample_range(header_row, len(selected_rows))
    if source_sheet.auto_filter and source_sheet.auto_filter.ref:
        if output_sheet.auto_filter.ref != expected_ref:
            raise OutputValidationError("输出自动筛选范围未正确更新")
    elif output_sheet.auto_filter and output_sheet.auto_filter.ref:
        raise OutputValidationError("输出文件出现源文件不存在的自动筛选范围")

    if set(source_sheet.tables.keys()) != set(output_sheet.tables.keys()):
        raise OutputValidationError("输出 Excel 表格对象与源文件不一致")
    for table_name in source_sheet.tables.keys():
        if output_sheet.tables[table_name].ref != expected_ref:
            raise OutputValidationError("输出 Excel 表格对象范围未正确更新")


def _remove_temp_file(path: Path) -> None:
    try:
        path.unlink(missing_ok=True)
    except OSError:
        pass
