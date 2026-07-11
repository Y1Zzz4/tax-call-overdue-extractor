"""50条样本批量提取、标准化、Excel回填与断点续跑。"""

from __future__ import annotations

import asyncio
import hashlib
import json
import logging
import os
import shutil
import time
import uuid
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from tax_call_overdue_extractor.config import ProjectSettings
from tax_call_overdue_extractor.excel_io import CALL_TEXT_COLUMN, EXPECTED_COLUMNS, hash_file, validate_header
from tax_call_overdue_extractor.exceptions import ExcelIOError, ExtractionError, LLMClientError, ResponseParseError
from tax_call_overdue_extractor.llm.client import LLMClientProtocol, OpenAICompatibleLLMClient
from tax_call_overdue_extractor.llm.request_builder import BUSINESS_CONTENT_KEY, REPLY_CONTENT_KEY, build_chat_request, build_model_input, load_system_prompt

from .batch_models import BatchPlan, BatchRecord, BatchRunSummary, BatchStatus, NormalizedItem, RowBatchOutcome, TextStats
from .normalization import normalize_extraction_result, parse_reference_month
from .parser import parse_extraction_response
from .schemas import ExtractionResult, STANDARD_TAX_TYPES, no_text_result
from .state_store import BatchStateStore


LOGGER = logging.getLogger(__name__)
SCHEMA_VERSION = "1.0"
SAFE_SAMPLE_LIMIT = 100
RESULT_COLUMNS = {
    "企业名称": 12,
    "逾期税种": 13,
    "所属期": 14,
    "涉及金额": 15,
    "是否确定已逾期": 16,
    "说明": 17,
}
UNKNOWN_ENTERPRISE = "未识别"
UNKNOWN_TAX = "未识别"
UNKNOWN_PERIOD = "未识别"
UNKNOWN_AMOUNT = "未提及"
UNKNOWN_OVERDUE = "未明确"


@dataclass(frozen=True)
class BatchOptions:
    input_path: Path | None = None
    output_path: Path | None = None
    conflicts_output_path: Path | None = None
    review_output_path: Path | None = None
    state_db_path: Path | None = None
    sheet_name: str | None = None
    rows: tuple[int, ...] | None = None
    max_records: int | None = None
    concurrency: int | None = None
    resume: bool = False
    execute: bool = False
    overwrite: bool = False
    allow_large_run: bool = False


class BatchExtractionService:
    """批处理服务，默认只预检，显式 execute 才调用模型。"""

    def __init__(
        self,
        settings: ProjectSettings,
        *,
        llm_client: LLMClientProtocol | None = None,
        system_prompt: str | None = None,
    ) -> None:
        self._settings = settings
        self._llm_client = llm_client
        self._system_prompt = system_prompt

    def build_plan(self, options: BatchOptions) -> tuple[BatchPlan, list[BatchRecord], str, str]:
        source_path = _resolve_batch_input_file(options.input_path, self._settings.paths.samples_dir)
        output_path = options.output_path or self._settings.paths.output_dir / f"{source_path.stem}_extracted.xlsx"
        conflicts_output = options.conflicts_output_path or self._settings.paths.conflicts_dir / f"{source_path.stem}_conflicts.xlsx"
        review_output = options.review_output_path or self._settings.paths.output_dir / f"{source_path.stem}_review.xlsx"
        state_db = options.state_db_path or self._settings.paths.state_dir / "batch_state.sqlite3"
        concurrency = options.concurrency or self._settings.llm_reserved.max_concurrency or 2
        prompt = self._system_prompt if self._system_prompt is not None else load_system_prompt()
        prompt_hash = _sha256_text(prompt)
        source_fingerprint = hash_file(source_path)

        records, sheet_name, total_rows = read_batch_records(
            source_path,
            sheet_name=options.sheet_name or self._settings.excel.sheet_name,
            use_active_sheet=self._settings.excel.use_active_sheet_when_sheet_not_set,
            header_row=self._settings.excel.header_row,
            rows=options.rows,
            max_records=options.max_records,
        )

        reusable = 0
        if options.resume:
            store = BatchStateStore(state_db)
            try:
                for record in records:
                    if store.reusable_record(
                        source_file_fingerprint=source_fingerprint,
                        worksheet=sheet_name,
                        original_row_number=record.original_row_number,
                        input_hash=record.input_hash,
                        prompt_hash=prompt_hash,
                        schema_version=SCHEMA_VERSION,
                        model_name=self._settings.llm.model,
                    ):
                        reusable += 1
            finally:
                store.close()

        eligible = sum(1 for record in records if record.model_input.has_any_text)
        too_long = sum(1 for record in records if record.model_input.total_chars > self._settings.llm.max_input_chars)
        estimated_api = sum(
            1
            for record in records
            if record.model_input.has_any_text and record.model_input.total_chars <= self._settings.llm.max_input_chars
        )
        if options.resume:
            estimated_api = max(0, estimated_api - reusable)

        plan = BatchPlan(
            input_path=source_path,
            output_path=output_path,
            conflicts_output_path=conflicts_output,
            review_output_path=review_output,
            state_db_path=state_db,
            sheet_name=sheet_name,
            original_data_rows=total_rows,
            eligible_rows=eligible,
            planned_records=len(records),
            reusable_records=reusable,
            estimated_api_calls=estimated_api,
            input_too_long_records=too_long,
            text_stats=_text_stats([record.model_input.total_chars for record in records]),
            model_name=self._settings.llm.model,
            concurrency=concurrency,
        )
        return plan, records, source_fingerprint, prompt_hash

    def preflight(self, options: BatchOptions) -> BatchPlan:
        plan, _, _, _ = self.build_plan(options)
        return plan

    def run(self, options: BatchOptions) -> BatchRunSummary:
        if not options.execute:
            raise ExtractionError("批处理默认只预检；必须显式传入 --execute 才会调用模型")

        plan, records, source_fingerprint, prompt_hash = self.build_plan(options)
        if plan.planned_records > SAFE_SAMPLE_LIMIT and not options.allow_large_run:
            raise ExtractionError(f"本次计划处理 {plan.planned_records} 条，超过样本阶段安全上限 {SAFE_SAMPLE_LIMIT} 条")
        if plan.output_path.exists() and not options.overwrite:
            raise ExtractionError(f"输出文件已存在，默认拒绝覆盖: {plan.output_path}")

        client = self._llm_client or OpenAICompatibleLLMClient(self._settings.llm)
        store = BatchStateStore(plan.state_db_path)
        run_id = _new_run_id()
        try:
            outcomes = asyncio.run(
                self._run_async(
                    records=records,
                    plan=plan,
                    client=client,
                    store=store,
                    source_fingerprint=source_fingerprint,
                    prompt_hash=prompt_hash,
                    resume=options.resume,
                    run_id=run_id,
                )
            )
        finally:
            store.close()

        write_outputs(
            input_path=plan.input_path,
            output_path=plan.output_path,
            conflicts_output_path=plan.conflicts_output_path,
            review_output_path=plan.review_output_path,
            sheet_name=plan.sheet_name,
            header_row=self._settings.excel.header_row,
            outcomes=outcomes,
            overwrite=options.overwrite,
        )
        return _summary(plan, outcomes)

    async def _run_async(
        self,
        *,
        records: list[BatchRecord],
        plan: BatchPlan,
        client: LLMClientProtocol,
        store: BatchStateStore,
        source_fingerprint: str,
        prompt_hash: str,
        resume: bool,
        run_id: str,
    ) -> list[RowBatchOutcome]:
        semaphore = asyncio.Semaphore(plan.concurrency)
        tasks = [
            self._process_record(
                record=record,
                plan=plan,
                client=client,
                store=store,
                source_fingerprint=source_fingerprint,
                prompt_hash=prompt_hash,
                resume=resume,
                run_id=run_id,
                semaphore=semaphore,
            )
            for record in records
        ]
        outcomes = await asyncio.gather(*tasks)
        return sorted(outcomes, key=lambda outcome: outcome.record.original_row_number)

    async def _process_record(
        self,
        *,
        record: BatchRecord,
        plan: BatchPlan,
        client: LLMClientProtocol,
        store: BatchStateStore,
        source_fingerprint: str,
        prompt_hash: str,
        resume: bool,
        run_id: str,
        semaphore: asyncio.Semaphore,
    ) -> RowBatchOutcome:
        reusable = None
        if resume:
            reusable = store.reusable_record(
                source_file_fingerprint=source_fingerprint,
                worksheet=plan.sheet_name,
                original_row_number=record.original_row_number,
                input_hash=record.input_hash,
                prompt_hash=prompt_hash,
                schema_version=SCHEMA_VERSION,
                model_name=self._settings.llm.model,
            )
        if reusable is not None:
            return _outcome_from_reusable(record, reusable)

        if not record.model_input.has_any_text:
            result = no_text_result()
            path = _save_structured_result(self._settings.paths.state_dir, run_id, record.original_row_number, result, "skipped_no_text")
            store.upsert(
                source_file_fingerprint=source_fingerprint,
                worksheet=plan.sheet_name,
                original_row_number=record.original_row_number,
                input_hash=record.input_hash,
                prompt_hash=prompt_hash,
                schema_version=SCHEMA_VERSION,
                model_name=self._settings.llm.model,
                status="skipped_no_text",
                attempts=0,
                structured_result_path=path,
                raw_response_path=None,
                error_type=None,
                error_message_sanitized=None,
            )
            return RowBatchOutcome(record, "skipped_no_text", result, [], path, None)

        if record.model_input.total_chars > self._settings.llm.max_input_chars:
            store.upsert(
                source_file_fingerprint=source_fingerprint,
                worksheet=plan.sheet_name,
                original_row_number=record.original_row_number,
                input_hash=record.input_hash,
                prompt_hash=prompt_hash,
                schema_version=SCHEMA_VERSION,
                model_name=self._settings.llm.model,
                status="input_too_long",
                attempts=0,
                structured_result_path=None,
                raw_response_path=None,
                error_type="input_too_long",
                error_message_sanitized=f"total_chars={record.model_input.total_chars}",
            )
            return RowBatchOutcome(record, "input_too_long", None, [], None, None, "input_too_long", "input_too_long")

        async with semaphore:
            attempts = store.mark_processing(
                source_file_fingerprint=source_fingerprint,
                worksheet=plan.sheet_name,
                original_row_number=record.original_row_number,
                input_hash=record.input_hash,
                prompt_hash=prompt_hash,
                schema_version=SCHEMA_VERSION,
                model_name=self._settings.llm.model,
            )
            raw_path: Path | None = None
            try:
                request = build_chat_request(self._system_prompt or load_system_prompt(), record.model_input)
                if self._llm_client is not None:
                    response = client.complete(request)
                else:
                    response = await asyncio.to_thread(client.complete, request)
                raw_path = _save_raw_response(self._settings.paths.state_dir, run_id, record.original_row_number, response.content)
                result = parse_extraction_response(response.content, dict(record.model_input.data))
                normalized = normalize_extraction_result(result, reference_month=record.reference_month)
                status = _status_for_result(result, normalized)
                structured_path = _save_structured_result(self._settings.paths.state_dir, run_id, record.original_row_number, result, status)
                store.upsert(
                    source_file_fingerprint=source_fingerprint,
                    worksheet=plan.sheet_name,
                    original_row_number=record.original_row_number,
                    input_hash=record.input_hash,
                    prompt_hash=prompt_hash,
                    schema_version=SCHEMA_VERSION,
                    model_name=self._settings.llm.model,
                    status=status,
                    attempts=attempts,
                    structured_result_path=structured_path,
                    raw_response_path=raw_path,
                    error_type=None,
                    error_message_sanitized=None,
                )
                return RowBatchOutcome(record, status, result, normalized, structured_path, raw_path)
            except ResponseParseError as exc:
                store.upsert(
                    source_file_fingerprint=source_fingerprint,
                    worksheet=plan.sheet_name,
                    original_row_number=record.original_row_number,
                    input_hash=record.input_hash,
                    prompt_hash=prompt_hash,
                    schema_version=SCHEMA_VERSION,
                    model_name=self._settings.llm.model,
                    status="validation_error",
                    attempts=attempts,
                    structured_result_path=None,
                    raw_response_path=raw_path,
                    error_type=exc.__class__.__name__,
                    error_message_sanitized=str(exc),
                )
                return RowBatchOutcome(record, "validation_error", None, [], None, raw_path, exc.__class__.__name__, str(exc))
            except LLMClientError as exc:
                store.upsert(
                    source_file_fingerprint=source_fingerprint,
                    worksheet=plan.sheet_name,
                    original_row_number=record.original_row_number,
                    input_hash=record.input_hash,
                    prompt_hash=prompt_hash,
                    schema_version=SCHEMA_VERSION,
                    model_name=self._settings.llm.model,
                    status="api_error",
                    attempts=attempts,
                    structured_result_path=None,
                    raw_response_path=None,
                    error_type=exc.__class__.__name__,
                    error_message_sanitized=str(exc),
                )
                return RowBatchOutcome(record, "api_error", None, [], None, None, exc.__class__.__name__, str(exc))


def read_batch_records(
    input_path: Path,
    *,
    sheet_name: str | None,
    use_active_sheet: bool,
    header_row: int,
    rows: tuple[int, ...] | None,
    max_records: int | None,
) -> tuple[list[BatchRecord], str, int]:
    workbook = load_workbook(input_path)
    try:
        if sheet_name:
            if sheet_name not in workbook.sheetnames:
                raise ExcelIOError(f"找不到指定工作表: {sheet_name}")
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.active if use_active_sheet else workbook[workbook.sheetnames[0]]
        header = validate_header(worksheet, header_row)
        column_map = {name: index + 1 for index, name in enumerate(header)}
        data_rows = list(range(header_row + 1, worksheet.max_row + 1))
        if rows is not None:
            row_set = set(rows)
            data_rows = [row for row in data_rows if row in row_set]
        if max_records is not None:
            data_rows = data_rows[:max_records]
        records: list[BatchRecord] = []
        for row_number in data_rows:
            model_input = build_model_input(
                voice_text=worksheet.cell(row=row_number, column=column_map[CALL_TEXT_COLUMN]).value,
                business_content=worksheet.cell(row=row_number, column=column_map[BUSINESS_CONTENT_KEY]).value,
                reply_content=worksheet.cell(row=row_number, column=column_map[REPLY_CONTENT_KEY]).value,
            )
            reference = parse_reference_month(
                worksheet.cell(row=row_number, column=column_map["登记日期"]).value,
                worksheet.cell(row=row_number, column=column_map["月份"]).value,
            )
            records.append(
                BatchRecord(
                    original_row_number=row_number,
                    sequence_number=worksheet.cell(row=row_number, column=1).value,
                    business_id=worksheet.cell(row=row_number, column=2).value,
                    model_input=model_input,
                    input_hash=model_input.sha256,
                    reference_month=reference,
                )
            )
        return records, worksheet.title, worksheet.max_row - header_row
    finally:
        workbook.close()


def write_outputs(
    *,
    input_path: Path,
    output_path: Path,
    conflicts_output_path: Path,
    review_output_path: Path,
    sheet_name: str,
    header_row: int,
    outcomes: list[RowBatchOutcome],
    overwrite: bool,
) -> None:
    if output_path.exists() and not overwrite:
        raise ExtractionError(f"输出文件已存在，默认拒绝覆盖: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.parent / f".{output_path.stem}.{uuid.uuid4().hex}.tmp.xlsx"
    source_hash = hash_file(input_path)
    try:
        shutil.copy2(input_path, temp_path)
        workbook = load_workbook(temp_path)
        try:
            worksheet = workbook[sheet_name]
            validate_header(worksheet, header_row)
            _write_result_rows(worksheet, outcomes, header_row)
            workbook.save(temp_path)
        finally:
            workbook.close()
        _validate_output_workbook(input_path, temp_path, sheet_name, header_row, outcomes)
        if hash_file(input_path) != source_hash:
            raise ExtractionError("原始输入文件在批处理过程中发生变化，已放弃输出")
        os.replace(temp_path, output_path)
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise

    _write_conflicts(conflicts_output_path, outcomes)
    _write_review(review_output_path, outcomes)


def _write_result_rows(worksheet, outcomes: list[RowBatchOutcome], header_row: int) -> None:
    explanation_header = worksheet.cell(row=header_row, column=RESULT_COLUMNS["说明"])
    explanation_header.value = "说明"
    source_header = worksheet.cell(row=header_row, column=16)
    explanation_header.font = copy(source_header.font)
    explanation_header.fill = copy(source_header.fill)
    explanation_header.border = copy(source_header.border)
    explanation_header.alignment = copy(source_header.alignment)
    for outcome in sorted(outcomes, key=lambda value: value.record.original_row_number):
        row_number = outcome.record.original_row_number
        if outcome.normalized_items:
            _write_item(worksheet, row_number, _merge_normalized_items(outcome.normalized_items))
        else:
            _write_empty_outcome(worksheet, row_number, outcome)


def _write_item(worksheet, row_number: int, item: NormalizedItem) -> None:
    worksheet.cell(row=row_number, column=12).value = item.enterprise_name or UNKNOWN_ENTERPRISE
    worksheet.cell(row=row_number, column=13).value = item.tax_types_text or UNKNOWN_TAX
    worksheet.cell(row=row_number, column=14).value = item.periods_text or UNKNOWN_PERIOD
    worksheet.cell(row=row_number, column=15).value = item.amounts_text or UNKNOWN_AMOUNT
    worksheet.cell(row=row_number, column=16).value = item.overdue_text or UNKNOWN_OVERDUE
    worksheet.cell(row=row_number, column=17).value = item.explanation


def _write_empty_outcome(worksheet, row_number: int, outcome: RowBatchOutcome) -> None:
    values = [UNKNOWN_ENTERPRISE, UNKNOWN_TAX, UNKNOWN_PERIOD, UNKNOWN_AMOUNT, UNKNOWN_OVERDUE]
    for column, value in enumerate(values, start=12):
        worksheet.cell(row=row_number, column=column).value = value
    explanations = {
        "skipped_no_text": "三列均无可分析文本",
        "input_too_long": "文本过长，模型未处理",
        "api_error": "模型调用失败，等待重试",
        "validation_error": "模型响应无法解析，等待重试",
    }
    worksheet.cell(row=row_number, column=17).value = explanations.get(
        outcome.status, "未发现税款或申报逾期信息"
    )


def _merge_normalized_items(items: list[NormalizedItem]) -> NormalizedItem:
    def merge_text(values: list[str | None]) -> str | None:
        parts: list[str] = []
        for value in values:
            if value:
                parts.extend(part for part in value.split("；") if part)
        unique = list(dict.fromkeys(parts))
        return "；".join(unique) if unique else None

    overdue_values = [item.overdue_text for item in items]
    if "已逾期" in overdue_values:
        overdue = "已逾期"
    elif overdue_values and all(value == "未逾期" for value in overdue_values):
        overdue = "未逾期"
    else:
        overdue = None
    return NormalizedItem(
        enterprise_name=merge_text([item.enterprise_name for item in items]),
        tax_types_text=merge_text([item.tax_types_text for item in items]),
        periods_text=merge_text([item.periods_text for item in items]),
        amounts_text=merge_text([item.amounts_text for item in items]),
        overdue_text=overdue,
        explanation=merge_text([item.explanation for item in items]) or "已完成提取",
        needs_review=any(item.needs_review for item in items),
        review_reasons=list(dict.fromkeys(reason for item in items for reason in item.review_reasons)),
        conflicts=[conflict for item in items for conflict in item.conflicts],
    )


def _write_conflicts(path: Path, outcomes: list[RowBatchOutcome]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "conflicts"
    headers = ["原Excel行号", "序号", "业务编号", "冲突字段", "冲突说明", "电话录音转文本内容的主张", "业务内容的主张", "答复内容的主张", "各来源证据", "处理方式", "是否需要人工复核"]
    worksheet.append(headers)
    _style_header(worksheet)
    for outcome in outcomes:
        for item in outcome.normalized_items:
            for conflict in item.conflicts:
                claims = {claim.source: claim for claim in conflict.claims}
                worksheet.append(
                    [
                        outcome.record.original_row_number,
                        outcome.record.sequence_number,
                        outcome.record.business_id,
                        conflict.field,
                        conflict.description,
                        claims.get("电话录音转文本内容").value if claims.get("电话录音转文本内容") else None,
                        claims.get("业务内容").value if claims.get("业务内容") else None,
                        claims.get("答复内容").value if claims.get("答复内容") else None,
                        CHINESE_JOIN([claim.quote for claim in conflict.claims]),
                        "主结果冲突字段保持为空，人工复核",
                        "是",
                    ]
                )
    workbook.save(path)
    workbook.close()


def _write_review(path: Path, outcomes: list[RowBatchOutcome]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "review"
    worksheet.append(["原Excel行号", "序号", "业务编号", "问题类型", "问题说明", "是否成功提取", "结构化结果文件路径", "建议处理方式"])
    _style_header(worksheet)
    for outcome in outcomes:
        if outcome.status in {"api_error", "validation_error", "input_too_long"}:
            issue = outcome.status if outcome.status != "input_too_long" else "input_too_long"
            worksheet.append([outcome.record.original_row_number, outcome.record.sequence_number, outcome.record.business_id, issue, outcome.error_message_sanitized, "否", outcome.structured_result_path and str(outcome.structured_result_path), "人工处理或重试"])
        if outcome.status == "skipped_no_text":
            worksheet.append([outcome.record.original_row_number, outcome.record.sequence_number, outcome.record.business_id, "no_extractable_information", "三列均无可分析文本", "否", outcome.structured_result_path and str(outcome.structured_result_path), "无需调用模型"])
        if outcome.status == "conflict":
            worksheet.append([outcome.record.original_row_number, outcome.record.sequence_number, outcome.record.business_id, "clear_enterprise_conflict", "答复内容与其他来源的企业名称明确不一致", "是", outcome.structured_result_path and str(outcome.structured_result_path), "人工确认企业名称"])
        for item in outcome.normalized_items:
            for reason in item.review_reasons if item.needs_review else []:
                worksheet.append([outcome.record.original_row_number, outcome.record.sequence_number, outcome.record.business_id, reason, reason, "是", outcome.structured_result_path and str(outcome.structured_result_path), "人工复核"])
    workbook.save(path)
    workbook.close()


def _validate_output_workbook(input_path: Path, output_path: Path, sheet_name: str, header_row: int, outcomes: list[RowBatchOutcome]) -> None:
    source = load_workbook(input_path)
    output = load_workbook(output_path)
    try:
        source_sheet = source[sheet_name]
        output_sheet = output[sheet_name]
        output_header = tuple(output_sheet.cell(row=header_row, column=column).value for column in range(1, 17))
        if output_header != EXPECTED_COLUMNS or output_sheet.cell(row=header_row, column=17).value != "说明":
            raise ExtractionError("输出表头不正确")
        if output_sheet.max_row != source_sheet.max_row:
            raise ExtractionError("输出不得增加、删除或拆分原始数据行")
        for row in range(header_row + 1, source_sheet.max_row + 1):
            for col in range(1, 12):
                source_cell = source_sheet.cell(row=row, column=col)
                output_cell = output_sheet.cell(row=row, column=col)
                if source_cell.value != output_cell.value:
                    raise ExtractionError("输出原始记录前11列发生变化")
                if source_cell.data_type != output_cell.data_type:
                    raise ExtractionError("输出原始记录前11列数据类型发生变化")
                if source_cell.number_format != output_cell.number_format:
                    raise ExtractionError("输出原始记录前11列显示格式发生变化")
                if source_cell._style != output_cell._style:
                    raise ExtractionError("输出原始记录前11列单元格样式发生变化")
        for row in range(header_row + 1, output_sheet.max_row + 1):
            tax_value = output_sheet.cell(row=row, column=13).value
            overdue_value = output_sheet.cell(row=row, column=16).value
            if tax_value:
                for tax_type in str(tax_value).split("；"):
                    if tax_type not in EXPECTED_TAX_TYPES:
                        raise ExtractionError("输出税种不在规范列表中")
            if overdue_value not in {None, "已逾期", "未逾期", "未明确"}:
                raise ExtractionError("逾期列只能是已逾期、未逾期或未明确")
    finally:
        source.close()
        output.close()


def _outcome_from_reusable(record: BatchRecord, reusable) -> RowBatchOutcome:
    result = None
    normalized: list[NormalizedItem] = []
    if reusable.structured_result_path and reusable.structured_result_path.exists():
        payload = json.loads(reusable.structured_result_path.read_text(encoding="utf-8"))
        result = ExtractionResult.model_validate(payload["result"])
        normalized = normalize_extraction_result(result, reference_month=record.reference_month)
    return RowBatchOutcome(record, reusable.status, result, normalized, reusable.structured_result_path, reusable.raw_response_path)


def _status_for_result(result: ExtractionResult, normalized: list[NormalizedItem]) -> BatchStatus:
    if not result.has_relevant_information:
        return "needs_review" if result.needs_review else "success"
    if any(item.conflicts for item in normalized) or result.conflicts:
        return "conflict"
    if result.needs_review or any(item.needs_review for item in normalized):
        return "needs_review"
    return "success"


def _new_run_id() -> str:
    return time.strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:6]


def _row_run_directory(state_dir: Path, run_id: str, row_number: int) -> Path:
    return state_dir / "runs" / run_id / f"row_{row_number:06d}"


def _save_raw_response(state_dir: Path, run_id: str, row_number: int, content: str) -> Path:
    directory = _row_run_directory(state_dir, run_id, row_number)
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / "response.txt"
    path.write_text(content, encoding="utf-8")
    return path


def _save_structured_result(state_dir: Path, run_id: str, row_number: int, result: ExtractionResult, status: str) -> Path:
    directory = _row_run_directory(state_dir, run_id, row_number)
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / "result.json"
    path.write_text(json.dumps({"status": status, "result": result.model_dump(mode="json")}, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def _resolve_batch_input_file(input_path: Path | None, samples_dir: Path) -> Path:
    if input_path is not None:
        if not input_path.exists():
            raise ExcelIOError(f"输入文件不存在: {input_path}")
        if input_path.suffix.lower() != ".xlsx":
            raise ExcelIOError(f"输入文件不是 .xlsx 文件: {input_path}")
        return input_path
    files = sorted(path for path in samples_dir.glob("*.xlsx") if path.is_file() and not path.name.startswith("~$"))
    if not files:
        raise ExcelIOError(f"样本目录下没有 .xlsx 文件: {samples_dir}")
    if len(files) > 1:
        raise ExcelIOError(f"样本目录下存在多个 .xlsx 文件，请使用 --input 指定: {samples_dir}")
    return files[0]


def _summary(plan: BatchPlan, outcomes: list[RowBatchOutcome]) -> BatchRunSummary:
    return BatchRunSummary(
        plan=plan,
        success_count=sum(1 for outcome in outcomes if outcome.status == "success"),
        conflict_count=sum(1 for outcome in outcomes if outcome.status == "conflict"),
        needs_review_count=sum(1 for outcome in outcomes if outcome.status == "needs_review"),
        skipped_count=sum(1 for outcome in outcomes if outcome.status == "skipped_no_text"),
        input_too_long_count=sum(1 for outcome in outcomes if outcome.status == "input_too_long"),
        api_error_count=sum(1 for outcome in outcomes if outcome.status == "api_error"),
        validation_error_count=sum(1 for outcome in outcomes if outcome.status == "validation_error"),
        output_path=plan.output_path,
        conflicts_output_path=plan.conflicts_output_path,
        review_output_path=plan.review_output_path,
        state_db_path=plan.state_db_path,
    )


def _text_stats(values: Sequence[int]) -> TextStats:
    if not values:
        return TextStats(total_chars=0, min_chars=0, max_chars=0, avg_chars=0)
    return TextStats(total_chars=sum(values), min_chars=min(values), max_chars=max(values), avg_chars=sum(values) / len(values))


def _sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def _style_header(worksheet) -> None:
    for cell in worksheet[1]:
        cell.font = Font(bold=True)


def CHINESE_JOIN(values: list[str]) -> str:
    return "；".join(values)


EXPECTED_TAX_TYPES = set(STANDARD_TAX_TYPES)
