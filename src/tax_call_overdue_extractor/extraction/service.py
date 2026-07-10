"""单条记录结构化提取服务。"""

from __future__ import annotations

import json
import logging
import os
import time
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Literal

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile

from tax_call_overdue_extractor.config import ProjectSettings
from tax_call_overdue_extractor.excel_io import CALL_TEXT_COLUMN, EXPECTED_COLUMNS, validate_header
from tax_call_overdue_extractor.exceptions import ExcelIOError, ExtractionError
from tax_call_overdue_extractor.llm.client import LLMClientProtocol, OpenAICompatibleLLMClient
from tax_call_overdue_extractor.llm.request_builder import (
    BUSINESS_CONTENT_KEY,
    REPLY_CONTENT_KEY,
    ChatRequest,
    FieldSummary,
    ModelInput,
    build_chat_request,
    build_model_input,
    load_system_prompt,
)

from .parser import parse_extraction_response
from .schemas import ExtractionResult, no_text_result


LOGGER = logging.getLogger(__name__)
RunStatus = Literal["success", "no_text", "input_too_long"]


@dataclass(frozen=True)
class DryRunSummary:
    row_number: int
    field_summaries: tuple[FieldSummary, ...]
    total_chars: int
    request_sha256: str
    model: str
    base_url: str


@dataclass(frozen=True)
class ExtractOneRunResult:
    row_number: int
    status: RunStatus
    called_api: bool
    item_count: int
    conflict_count: int
    needs_review: bool
    output_path: Path
    duration_seconds: float


class SingleRecordExtractionService:
    """面向 CLI 的单条记录提取编排服务。"""

    def __init__(
        self,
        settings: ProjectSettings,
        *,
        llm_client: LLMClientProtocol | None = None,
        system_prompt: str | None = None,
    ) -> None:
        self._settings = settings
        self._llm_client = llm_client
        self._system_prompt = system_prompt

    def build_dry_run(
        self,
        *,
        input_path: str | Path | None,
        row_number: int,
        sheet_name: str | None,
    ) -> DryRunSummary:
        """构建 dry-run 摘要，不调用 API，不返回任何原文。"""

        source_path = resolve_extract_input_file(input_path, self._settings.paths.samples_dir)
        model_input = read_model_input_from_excel(
            source_path,
            row_number=row_number,
            sheet_name=sheet_name or self._settings.excel.sheet_name,
            use_active_sheet=self._settings.excel.use_active_sheet_when_sheet_not_set,
            header_row=self._settings.excel.header_row,
        )
        return DryRunSummary(
            row_number=row_number,
            field_summaries=model_input.field_summaries,
            total_chars=model_input.total_chars,
            request_sha256=model_input.sha256,
            model=self._settings.llm.model,
            base_url=self._settings.llm.base_url,
        )

    def extract_one(
        self,
        *,
        input_path: str | Path | None,
        row_number: int,
        output_path: str | Path | None,
        sheet_name: str | None,
        overwrite: bool,
    ) -> ExtractOneRunResult:
        """读取单行并执行一次结构化提取；本阶段不修改 Excel。"""

        started = time.perf_counter()
        source_path = resolve_extract_input_file(input_path, self._settings.paths.samples_dir)
        destination = _default_output_path(self._settings.paths.state_dir, row_number, output_path)
        _validate_output_path(destination, overwrite)

        model_input = read_model_input_from_excel(
            source_path,
            row_number=row_number,
            sheet_name=sheet_name or self._settings.excel.sheet_name,
            use_active_sheet=self._settings.excel.use_active_sheet_when_sheet_not_set,
            header_row=self._settings.excel.header_row,
        )
        if not model_input.has_any_text:
            result = no_text_result()
            _write_run_output(
                destination,
                status="no_text",
                row_number=row_number,
                called_api=False,
                model=self._settings.llm.model,
                base_url=self._settings.llm.base_url,
                model_input=model_input,
                result=result,
                raw_response_path=None,
            )
            return _summary(row_number, "no_text", False, result, destination, started)

        if model_input.total_chars > self._settings.llm.max_input_chars:
            _write_run_output(
                destination,
                status="input_too_long",
                row_number=row_number,
                called_api=False,
                model=self._settings.llm.model,
                base_url=self._settings.llm.base_url,
                model_input=model_input,
                result=None,
                raw_response_path=None,
                error_message=(
                    f"input_too_long: 输入总字符数 {model_input.total_chars} 超过上限 "
                    f"{self._settings.llm.max_input_chars}，后续需要使用分块提取流程"
                ),
            )
            LOGGER.warning(
                "单条提取输入过长 row=%s total_chars=%s max_input_chars=%s",
                row_number,
                model_input.total_chars,
                self._settings.llm.max_input_chars,
            )
            return ExtractOneRunResult(
                row_number=row_number,
                status="input_too_long",
                called_api=False,
                item_count=0,
                conflict_count=0,
                needs_review=True,
                output_path=destination,
                duration_seconds=time.perf_counter() - started,
            )

        request = self._build_chat_request(model_input)
        client = self._llm_client or OpenAICompatibleLLMClient(self._settings.llm)
        response = client.complete(request)
        raw_response_path = _save_raw_response(self._settings.paths.state_dir, row_number, response.content)
        result = parse_extraction_response(response.content)
        _write_run_output(
            destination,
            status="success",
            row_number=row_number,
            called_api=True,
            model=self._settings.llm.model,
            base_url=self._settings.llm.base_url,
            model_input=model_input,
            result=result,
            raw_response_path=raw_response_path,
        )
        return _summary(row_number, "success", True, result, destination, started)

    def _build_chat_request(self, model_input: ModelInput) -> ChatRequest:
        system_prompt = self._system_prompt if self._system_prompt is not None else load_system_prompt()
        return build_chat_request(system_prompt, model_input)


def resolve_extract_input_file(input_path: str | Path | None, samples_dir: Path) -> Path:
    """解析单条提取输入；未指定时要求 data/samples 下恰好有一个 .xlsx。"""

    if input_path is not None:
        path = Path(input_path)
        _validate_input_xlsx(path)
        return path

    if not samples_dir.exists() or not samples_dir.is_dir():
        raise ExcelIOError(f"抽样目录不存在或不可访问: {samples_dir}")
    files = sorted(
        path for path in samples_dir.glob("*.xlsx") if path.is_file() and not path.name.startswith("~$")
    )
    if not files:
        raise ExcelIOError(f"抽样目录下没有 .xlsx 文件: {samples_dir}")
    if len(files) > 1:
        raise ExcelIOError(f"抽样目录下存在多个 .xlsx 文件，请使用 --input 指定输入文件: {samples_dir}")
    _validate_input_xlsx(files[0])
    return files[0]


def read_model_input_from_excel(
    input_path: Path,
    *,
    row_number: int,
    sheet_name: str | None,
    use_active_sheet: bool,
    header_row: int,
) -> ModelInput:
    """从指定 Excel 行读取三个允许字段，并转换为模型输入。"""

    if row_number <= header_row:
        raise ExcelIOError(f"row-number 必须是数据行号，当前表头行为 {header_row}")
    _validate_input_xlsx(input_path)
    workbook = _load_workbook(input_path)
    try:
        worksheet = _select_sheet(workbook, sheet_name, use_active_sheet)
        header = validate_header(worksheet, header_row)
        if row_number > worksheet.max_row:
            raise ExcelIOError(f"指定行号超出工作表范围: {row_number}")
        column_map = {name: index + 1 for index, name in enumerate(header)}
        return build_model_input(
            voice_text=worksheet.cell(row=row_number, column=column_map[CALL_TEXT_COLUMN]).value,
            business_content=worksheet.cell(row=row_number, column=column_map[BUSINESS_CONTENT_KEY]).value,
            reply_content=worksheet.cell(row=row_number, column=column_map[REPLY_CONTENT_KEY]).value,
        )
    finally:
        workbook.close()


def _summary(
    row_number: int,
    status: RunStatus,
    called_api: bool,
    result: ExtractionResult,
    output_path: Path,
    started: float,
) -> ExtractOneRunResult:
    return ExtractOneRunResult(
        row_number=row_number,
        status=status,
        called_api=called_api,
        item_count=len(result.items),
        conflict_count=len(result.conflicts),
        needs_review=result.needs_review,
        output_path=output_path,
        duration_seconds=time.perf_counter() - started,
    )


def _write_run_output(
    output_path: Path,
    *,
    status: RunStatus,
    row_number: int,
    called_api: bool,
    model: str,
    base_url: str,
    model_input: ModelInput,
    result: ExtractionResult | None,
    raw_response_path: Path | None,
    error_message: str | None = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "status": status,
        "row_number": row_number,
        "called_api": called_api,
        "model": model,
        "base_url": base_url,
        "input_total_chars": model_input.total_chars,
        "request_sha256": model_input.sha256,
        "raw_response_path": str(raw_response_path) if raw_response_path is not None else None,
        "error_message": error_message,
        "result": result.model_dump(mode="json") if result is not None else None,
    }
    temp_path = output_path.parent / f".{output_path.stem}.{uuid.uuid4().hex}.tmp.json"
    try:
        temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(temp_path, output_path)
    except OSError as exc:
        temp_path.unlink(missing_ok=True)
        raise ExtractionError(f"写入单条提取结果失败: {output_path}") from exc


def _save_raw_response(state_dir: Path, row_number: int, content: str) -> Path:
    raw_dir = state_dir / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    path = raw_dir / f"row_{row_number}_{uuid.uuid4().hex}.txt"
    path.write_text(content, encoding="utf-8")
    return path


def _default_output_path(state_dir: Path, row_number: int, output_path: str | Path | None) -> Path:
    if output_path is not None:
        return Path(output_path)
    return state_dir / "preview" / f"row_{row_number}.json"


def _validate_output_path(output_path: Path, overwrite: bool) -> None:
    if output_path.exists() and not overwrite:
        raise ExtractionError(f"输出文件已存在，默认拒绝覆盖: {output_path}")


def _validate_input_xlsx(path: Path) -> None:
    if not path.exists():
        raise ExcelIOError(f"输入文件不存在: {path}")
    if not path.is_file():
        raise ExcelIOError(f"输入路径不是文件: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ExcelIOError(f"输入文件不是 .xlsx 文件: {path}")


def _load_workbook(path: Path):
    try:
        return load_workbook(path)
    except FileNotFoundError as exc:
        raise ExcelIOError(f"读取工作簿失败，文件不存在: {path}") from exc
    except PermissionError as exc:
        raise ExcelIOError(f"读取工作簿失败，文件无权限或被占用: {path}") from exc
    except (InvalidFileException, BadZipFile, OSError) as exc:
        raise ExcelIOError(f"读取工作簿失败，文件可能损坏或不可访问: {path}") from exc


def _select_sheet(workbook, sheet_name: str | None, use_active_sheet: bool):
    if sheet_name is not None:
        if sheet_name not in workbook.sheetnames:
            raise ExcelIOError(f"找不到指定工作表: {sheet_name}")
        return workbook[sheet_name]
    if use_active_sheet:
        return workbook.active
    if len(workbook.sheetnames) == 1:
        return workbook[workbook.sheetnames[0]]
    raise ExcelIOError("未指定工作表，且工作簿包含多个工作表")
